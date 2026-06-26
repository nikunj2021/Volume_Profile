"""
weekly_volume_profile.py
========================
Step 3 of 3  —  Resample daily OHLCV → weekly candles, build weekly volume
                profiles, cluster S/R zones, push Weekly SR Zones to Sheets.

Reads:   (yfinance directly — daily data for LOOKBACK_WEEKS weeks)
Outputs: volume_profile_weekly.xlsx
         Google Sheet tab → "Weekly SR Zones"

Columns in Weekly SR Zones (Google Sheet):
  A Symbol | B LTP (GOOGLEFINANCE live) | C Zone Price | D VAH | E VAL
  F Touches | G Type | H Strength | I % from LTP | J Signal | K Alert Sent

Signal formula (J) fires when LTP is within 0.5% of VAH or VAL.
Alert Sent (K) is written by oracle_alert.py after a Telegram is sent.

Run weekly on GitHub Actions (Saturday 8:00 AM IST / 02:30 UTC).
"""

import yfinance as yf
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

from config import (
    SYMBOLS,
    LOOKBACK_WEEKS, PRICE_BINS, VALUE_AREA_PCT,
    WEEKLY_TOLERANCE_PCT, WEEKLY_MIN_TOUCHES, WEEKLY_STRONG_TOUCHES, SR_TOP_N,
    WEEKLY_EXCEL,
    GOOGLE_SHEET_ID, GOOGLE_CREDS_JSON, SHEET_WEEKLY_SR,
    ALERT_BAND_PCT,
)

# ── Colour palette ─────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SEC_FILL = PatternFill("solid", fgColor="1F3864")
POC_FILL = PatternFill("solid", fgColor="FFF2CC")
VAH_FILL = PatternFill("solid", fgColor="E2EFDA")
VAL_FILL = PatternFill("solid", fgColor="FCE4D6")
COL_FILL = PatternFill("solid", fgColor="D9E1F2")
SR_FILLS = {
    "STRONG_R": PatternFill("solid", fgColor="C00000"),
    "STRONG_S": PatternFill("solid", fgColor="375623"),
    "R":        PatternFill("solid", fgColor="FF9999"),
    "S":        PatternFill("solid", fgColor="C6EFCE"),
}
SR_FONTS = {
    "STRONG_R": Font(bold=True, color="FFFFFF", name="Arial", size=9),
    "STRONG_S": Font(bold=True, color="FFFFFF", name="Arial", size=9),
    "R":        Font(bold=True, color="C00000", name="Arial", size=9),
    "S":        Font(bold=True, color="375623", name="Arial", size=9),
}


# ── Data fetch & resample ──────────────────────────────────────────────────────

def fetch_weekly_ohlcv(symbol: str) -> pd.DataFrame:
    ticker = symbol + ".NS"
    end    = datetime.today()
    start  = end - timedelta(days=LOOKBACK_WEEKS * 7 + 14)
    df = yf.download(ticker, start=start, end=end,
                     interval="1d", progress=False, auto_adjust=True)
    if df.empty:
        return df
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df = df.dropna()

    weekly = df.resample("W-FRI").agg({
        "Open":   "first",
        "High":   "max",
        "Low":    "min",
        "Close":  "last",
        "Volume": "sum",
    }).dropna()
    weekly = weekly[weekly["Volume"] > 0].tail(LOOKBACK_WEEKS)
    return weekly


# ── Profile engine ─────────────────────────────────────────────────────────────

def build_profile(row: pd.Series) -> dict | None:
    high, low, volume = float(row["High"]), float(row["Low"]), float(row["Volume"])
    if high == low or volume == 0:
        return None
    edges      = np.linspace(low, high, PRICE_BINS + 1)
    bin_prices = (edges[:-1] + edges[1:]) / 2
    bin_vols   = np.full(PRICE_BINS, volume / PRICE_BINS)
    poc        = float(bin_prices[int(np.argmax(bin_vols))])
    order      = np.argsort(bin_vols)[::-1]
    cumvol, va_idx = 0.0, []
    for i in order:
        cumvol += bin_vols[i]; va_idx.append(i)
        if cumvol / volume >= VALUE_AREA_PCT / 100:
            break
    va_px = bin_prices[va_idx]
    return {
        "bin_prices":   bin_prices.tolist(),
        "bin_volumes":  bin_vols.tolist(),
        "poc":          poc,
        "vah":          float(va_px.max()),
        "val":          float(va_px.min()),
        "total_volume": volume,
    }


# ── S/R clustering ─────────────────────────────────────────────────────────────

def cluster_levels(prices: list, tolerance_pct: float) -> list:
    if not prices:
        return []
    prices = sorted(prices)
    groups, group = [], [prices[0]]
    for p in prices[1:]:
        mean = sum(group) / len(group)
        if abs(p - mean) / mean * 100 <= tolerance_pct:
            group.append(p)
        else:
            groups.append(group); group = [p]
    groups.append(group)
    return [{"mean": sum(g)/len(g), "count": len(g), "prices": g} for g in groups]


def build_sr_zones(sessions: list) -> list:
    key_prices = [lvl for s in sessions
                  for lvl in [s["poc"], s["vah"], s["val"]] if lvl is not None]
    clusters = cluster_levels(key_prices, WEEKLY_TOLERANCE_PCT)
    zones    = [c for c in clusters if c["count"] >= WEEKLY_MIN_TOUCHES]
    zones.sort(key=lambda z: z["mean"], reverse=True)
    return zones


def pick_top_zones(zones: list, last_close: float):
    strong = [z for z in zones if z["count"] >= WEEKLY_STRONG_TOUCHES]
    res    = sorted([z for z in strong if z["mean"] >= last_close], key=lambda z: z["mean"])[:SR_TOP_N]
    sup    = sorted([z for z in strong if z["mean"] <  last_close], key=lambda z: z["mean"], reverse=True)[:SR_TOP_N]
    return res, sup


# ── Excel writers ──────────────────────────────────────────────────────────────

def write_weekly_sheet(wb, symbol: str, sessions: list, zones: list) -> None:
    sheet_name = f"{symbol}_W"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws        = wb.create_sheet(sheet_name)
    row_ptr   = 1
    last_close = sessions[-1]["close"]

    for sess in sessions:
        ohlcv   = sess["ohlcv"]
        profile = sess["profile"]

        ws.cell(row_ptr, 1, f"Week: {sess['week_label']}").fill = SEC_FILL
        ws.cell(row_ptr, 1).font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        hdr_txt = (f"O:{float(ohlcv['Open']):.2f}  H:{float(ohlcv['High']):.2f}  "
                   f"L:{float(ohlcv['Low']):.2f}  C:{float(ohlcv['Close']):.2f}  "
                   f"Vol:{int(float(ohlcv['Volume'])):,}")
        ws.cell(row_ptr, 2, hdr_txt).fill = SEC_FILL
        ws.cell(row_ptr, 2).font = Font(color="FFFFFF", name="Arial", size=9)
        ws.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=6)
        row_ptr += 1

        for col, hdr in enumerate(["Price Bin", "Volume", "% of Week", "Level", "S/R"], 1):
            c = ws.cell(row_ptr, col, hdr)
            c.font = Font(bold=True, name="Arial", size=9)
            c.fill = COL_FILL
            c.alignment = Alignment(horizontal="center")
        row_ptr += 1

        total_vol = profile["total_volume"]
        poc, vah, val = profile["poc"], profile["vah"], profile["val"]

        for price, vol in sorted(zip(profile["bin_prices"], profile["bin_volumes"]), reverse=True):
            pct   = vol / total_vol * 100 if total_vol else 0
            level = ("POC" if abs(price - poc) < 1e-6 else
                     "VAH" if abs(price - vah) < 1e-6 else
                     "VAL" if abs(price - val) < 1e-6 else "")
            fill  = (POC_FILL if level == "POC" else
                     VAH_FILL if level == "VAH" else
                     VAL_FILL if level == "VAL" else None)

            # S/R label
            sr = None
            for z in zones:
                if abs(price - z["mean"]) / z["mean"] * 100 <= WEEKLY_TOLERANCE_PCT:
                    above  = price >= last_close
                    strong = z["count"] >= WEEKLY_STRONG_TOUCHES
                    sr = ("STRONG_R" if strong and above else
                          "STRONG_S" if strong else
                          "R"        if above  else "S")
                    break

            ws.cell(row_ptr, 1, round(price, 2)).font = Font(name="Arial", size=9)
            ws.cell(row_ptr, 2, round(vol, 0)).font   = Font(name="Arial", size=9)
            ws.cell(row_ptr, 3, f"{pct:.1f}%").font   = Font(name="Arial", size=9)
            ws.cell(row_ptr, 4, level).font            = Font(bold=bool(level), name="Arial", size=9)
            if sr:
                ws.cell(row_ptr, 5, sr).fill = SR_FILLS[sr]
                ws.cell(row_ptr, 5).font     = SR_FONTS[sr]
                ws.cell(row_ptr, 5).alignment = Alignment(horizontal="center")
            for col in range(1, 5):
                ws.cell(row_ptr, col).alignment = Alignment(horizontal="center")
                if fill:
                    ws.cell(row_ptr, col).fill = fill
            row_ptr += 1

        row_ptr += 1

    for col, w in zip([1,2,3,4,5],[12,16,12,8,10]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A1"


def write_weekly_sr_sheet(wb, sr_data: list) -> list:
    """Create 'Weekly SR Zones' sheet. Returns export_rows for Sheets push."""
    name = "Weekly SR Zones"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    headers = ["Symbol", "Zone Price", "VAH", "VAL", "Touches", "Type", "Strength"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(1, col)
        c.fill = HDR_FILL
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.alignment = Alignment(horizontal="center")

    export_rows = [headers[:]]

    for symbol, zones, last_close, sessions in sr_data:
        last_vah = round(sessions[-1]["vah"], 2)
        last_val = round(sessions[-1]["val"], 2)
        res, sup = pick_top_zones(zones, last_close)
        for z in res + sup:
            price     = round(z["mean"], 2)
            zone_type = "Resistance" if price >= last_close else "Support"
            row_data  = [symbol, price, last_vah, last_val, z["count"], zone_type, "STRONG"]
            ws.append(row_data)
            export_rows.append(row_data[:])
            r        = ws.max_row
            fill_key = "STRONG_R" if zone_type == "Resistance" else "STRONG_S"
            for col in range(1, 8):
                ws.cell(r, col).font      = Font(name="Arial", size=9)
                ws.cell(r, col).alignment = Alignment(horizontal="center")
            ws.cell(r, 2).fill = SR_FILLS[fill_key]
            ws.cell(r, 6).fill = SR_FILLS[fill_key]

    for i, w in enumerate([12,12,10,10,10,14,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return export_rows


# ── Google Sheets push ─────────────────────────────────────────────────────────

def push_weekly_to_sheets(export_rows: list) -> None:
    """
    Push Weekly SR Zones to Google Sheets.
    Column layout:
      A Symbol | B LTP (live GOOGLEFINANCE) | C Zone Price | D VAH | E VAL
      F Touches | G Type | H Strength | I % from LTP (live) | J Signal | K Alert Sent

    Signal (J) = sheet formula — fires when LTP is within ALERT_BAND_PCT of VAH or VAL.
    Alert Sent (K) = written by oracle_alert.py after Telegram fires.
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = ["https://spreadsheets.google.com/feeds",
                  "https://www.googleapis.com/auth/drive"]
        creds  = Credentials.from_service_account_file(GOOGLE_CREDS_JSON, scopes=scopes)
        gc     = gspread.authorize(creds)
        sh     = gc.open_by_key(GOOGLE_SHEET_ID)

        try:
            ws = sh.worksheet(SHEET_WEEKLY_SR)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=SHEET_WEEKLY_SR, rows=500, cols=12)

        ws.clear()

        sheet_headers = [
            "Symbol", "LTP (Live)", "Zone Price", "VAH", "VAL",
            "Touches", "Type", "Strength", "% from LTP", "Signal", "Alert Sent"
        ]
        ws.update("A1", [sheet_headers], value_input_option="RAW")

        # data rows = export_rows[1:]
        # export cols: [Symbol, ZonePrice, VAH, VAL, Touches, Type, Strength]
        band = ALERT_BAND_PCT / 100
        data_rows = export_rows[1:]
        batch = []

        for i, r in enumerate(data_rows, start=2):
            symbol     = r[0]
            zone_price = r[1]
            vah        = r[2]
            val        = r[3]
            touches    = r[4]
            ztype      = r[5]
            strength   = r[6]
            gf_ticker  = f"NSE:{symbol}"

            ltp_f    = f'=IFERROR(GOOGLEFINANCE("{gf_ticker}","price"),"")'
            pct_f    = f'=IF(B{i}="","",ROUND((B{i}-C{i})/C{i}*100,2))'
            signal_f = (
                f'=IF(B{i}="","⏳ Waiting",'
                f'IF(ABS(B{i}-D{i})/D{i}<={band},"⚠️ Near VAH",'
                f'IF(ABS(B{i}-E{i})/E{i}<={band},"⚠️ Near VAL","✅ Clear")))'
            )

            batch.append({"range": f"A{i}:A{i}", "values": [[symbol]]})
            batch.append({"range": f"B{i}",      "values": [[ltp_f]]})
            batch.append({"range": f"C{i}:H{i}", "values": [[zone_price, vah, val, touches, ztype, strength]]})
            batch.append({"range": f"I{i}",      "values": [[pct_f]]})
            batch.append({"range": f"J{i}",      "values": [[signal_f]]})
            # K (Alert Sent) left blank — oracle_alert.py writes "YES" here

        ws.batch_update(batch, value_input_option="USER_ENTERED")

        # ── Formatting ────────────────────────────────────────────────────
        n = len(data_rows)
        requests = [
            {"repeatCell": {
                "range": {"sheetId": ws.id, "startRowIndex": 0, "endRowIndex": 1,
                          "startColumnIndex": 0, "endColumnIndex": 11},
                "cell": {"userEnteredFormat": {
                    "backgroundColor": {"red": 0.12, "green": 0.22, "blue": 0.39},
                    "textFormat": {"bold": True,
                                   "foregroundColor": {"red":1,"green":1,"blue":1}},
                    "horizontalAlignment": "CENTER",
                }},
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
            }},
            {"repeatCell": {
                "range": {"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": n+1,
                          "startColumnIndex": 0, "endColumnIndex": 11},
                "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER"}},
                "fields": "userEnteredFormat.horizontalAlignment",
            }},
            # Signal ⚠️ Near VAH → light red
            {"addConditionalFormatRule": {"rule": {
                "ranges": [{"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": n+1,
                            "startColumnIndex": 9, "endColumnIndex": 10}],
                "booleanRule": {
                    "condition": {"type": "TEXT_CONTAINS",
                                  "values": [{"userEnteredValue": "Near VAH"}]},
                    "format": {"backgroundColor": {"red":1.0,"green":0.8,"blue":0.8}},
                },
            }, "index": 0}},
            # Signal ⚠️ Near VAL → light green
            {"addConditionalFormatRule": {"rule": {
                "ranges": [{"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": n+1,
                            "startColumnIndex": 9, "endColumnIndex": 10}],
                "booleanRule": {
                    "condition": {"type": "TEXT_CONTAINS",
                                  "values": [{"userEnteredValue": "Near VAL"}]},
                    "format": {"backgroundColor": {"red":0.8,"green":1.0,"blue":0.85}},
                },
            }, "index": 1}},
            # Freeze row 1 + col A
            {"updateSheetProperties": {
                "properties": {"sheetId": ws.id,
                               "gridProperties": {"frozenRowCount": 1,
                                                  "frozenColumnCount": 1}},
                "fields": "gridProperties(frozenRowCount,frozenColumnCount)",
            }},
        ]
        for ci, px in enumerate([90, 100, 100, 80, 80, 80, 110, 90, 120, 140, 100]):
            requests.append({"updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS",
                          "startIndex": ci, "endIndex": ci+1},
                "properties": {"pixelSize": px},
                "fields": "pixelSize",
            }})

        sh.batch_update({"requests": requests})
        print(f"✅ Google Sheets '{SHEET_WEEKLY_SR}' updated → {n} rows")
        print(f"   Col B = GOOGLEFINANCE live LTP | Col J = proximity signal formula")
        print(f"   Alert band = ±{ALERT_BAND_PCT}% of VAH / VAL")

    except Exception as e:
        print(f"⚠ Google Sheets push failed: {e}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    try:
        wb = openpyxl.load_workbook(WEEKLY_EXCEL)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    sr_data = []

    for symbol in SYMBOLS:
        print(f"Processing {symbol} ...", flush=True)
        weekly_df = fetch_weekly_ohlcv(symbol)
        if weekly_df.empty:
            print(f"  ⚠ No data, skipping.")
            continue

        sessions = []
        for date, row in weekly_df.iterrows():
            profile = build_profile(row)
            if not profile:
                continue
            week_end   = pd.Timestamp(date)
            week_start = week_end - pd.Timedelta(days=4)
            sessions.append({
                "week_label": f"{week_start.strftime('%d %b')} – {week_end.strftime('%d %b %Y')}",
                "profile":    profile,
                "ohlcv":      row,
                "close":      float(row["Close"]),
                "poc":        profile["poc"],
                "vah":        profile["vah"],
                "val":        profile["val"],
            })

        if not sessions:
            continue

        zones      = build_sr_zones(sessions)
        last_close = sessions[-1]["close"]
        res, sup   = pick_top_zones(zones, last_close)
        print(f"  ✓ {len(sessions)} weeks | {len(res)} STRONG_R  {len(sup)} STRONG_S")

        write_weekly_sheet(wb, symbol, sessions, zones)
        sr_data.append((symbol, zones, last_close, sessions))

    export_rows = write_weekly_sr_sheet(wb, sr_data)
    wb.save(WEEKLY_EXCEL)
    print(f"\n✅ Saved → {WEEKLY_EXCEL}")

    if GOOGLE_SHEET_ID:
        push_weekly_to_sheets(export_rows)


if __name__ == "__main__":
    main()
