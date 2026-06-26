"""
volume_profile.py
=================
Step 1 of 3  —  Build daily session volume profiles for NSE stocks.

Outputs:  volume_profile_daily.xlsx
  • Summary sheet  — one row per symbol per session (POC / VAH / VAL)
  • One sheet per symbol — full price-bin breakdown for every session

Run weekly on GitHub Actions (Saturday 8:00 AM IST / 02:30 UTC).
"""

import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

from config import (
    SYMBOLS, LOOKBACK_DAYS, PRICE_BINS, VALUE_AREA_PCT, DAILY_EXCEL
)

# ── Colour palette ─────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SEC_FILL = PatternFill("solid", fgColor="2E4057")
POC_FILL = PatternFill("solid", fgColor="FFF2CC")
VAH_FILL = PatternFill("solid", fgColor="E2EFDA")
VAL_FILL = PatternFill("solid", fgColor="FCE4D6")
COL_FILL = PatternFill("solid", fgColor="D9E1F2")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)


# ── Data fetch ─────────────────────────────────────────────────────────────────

def fetch_daily_ohlcv(symbol: str, days: int) -> pd.DataFrame:
    ticker = symbol + ".NS"
    end    = datetime.today()
    start  = end - timedelta(days=days + 14)          # buffer for holidays
    df = yf.download(ticker, start=start, end=end,
                     interval="1d", progress=False, auto_adjust=True)
    if df.empty:
        return df
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    return df.dropna().tail(days)


# ── Profile engine ─────────────────────────────────────────────────────────────

def build_session_profile(row: pd.Series) -> dict | None:
    """
    Build volume profile for one daily candle.
    Volume is distributed uniformly across PRICE_BINS (standard EOD approximation).
    Returns dict with bin_prices, bin_volumes, poc, vah, val, total_volume — or None.
    """
    high   = float(row["High"])
    low    = float(row["Low"])
    volume = float(row["Volume"])
    if high == low or volume == 0:
        return None

    edges      = np.linspace(low, high, PRICE_BINS + 1)
    bin_prices = (edges[:-1] + edges[1:]) / 2          # midpoints
    bin_vols   = np.full(PRICE_BINS, volume / PRICE_BINS)

    # Point of Control
    poc = float(bin_prices[int(np.argmax(bin_vols))])

    # Value Area (70% of volume, highest-volume bins first)
    order      = np.argsort(bin_vols)[::-1]
    cumvol     = 0.0
    va_idx     = []
    for i in order:
        cumvol += bin_vols[i]
        va_idx.append(i)
        if cumvol / volume >= VALUE_AREA_PCT / 100:
            break
    va_px = bin_prices[va_idx]

    return {
        "bin_prices":   bin_prices.tolist(),
        "bin_volumes":  bin_vols.tolist(),
        "poc":          poc,
        "vah":          float(va_px.max()),
        "val":          float(va_px.min()),
        "total_volume": volume,
    }


# ── Excel writers ──────────────────────────────────────────────────────────────

def write_summary_sheet(wb: Workbook, rows: list) -> None:
    ws = wb.create_sheet("Summary", 0)
    headers = ["Symbol", "Date", "Open", "High", "Low", "Close",
               "Volume", "POC", "VAH", "VAL", "Range", "VA_Width"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(1, col)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, start=2):
        ws.append(row)
        ws.cell(ri, 8).fill  = POC_FILL
        ws.cell(ri, 9).fill  = VAH_FILL
        ws.cell(ri, 10).fill = VAL_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(ri, col).font      = Font(name="Arial", size=9)
            ws.cell(ri, col).alignment = Alignment(horizontal="center")

    for i, w in enumerate([12,12,10,10,10,10,14,10,10,10,10,10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_symbol_sheet(wb: Workbook, symbol: str, sessions: list) -> None:
    ws      = wb.create_sheet(symbol)
    row_ptr = 1

    for sess in sessions:
        ohlcv   = sess["ohlcv"]
        profile = sess["profile"]

        # Session header
        ws.cell(row_ptr, 1, f"Date: {sess['date']}").fill = SEC_FILL
        ws.cell(row_ptr, 1).font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        header_txt = (f"O:{float(ohlcv['Open']):.2f}  H:{float(ohlcv['High']):.2f}  "
                      f"L:{float(ohlcv['Low']):.2f}  C:{float(ohlcv['Close']):.2f}  "
                      f"Vol:{int(float(ohlcv['Volume'])):,}")
        ws.cell(row_ptr, 2, header_txt).fill = SEC_FILL
        ws.cell(row_ptr, 2).font = Font(color="FFFFFF", name="Arial", size=9)
        ws.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=5)
        row_ptr += 1

        # Column headers
        for col, hdr in enumerate(["Price Bin", "Volume", "% of Session", "Level"], 1):
            c = ws.cell(row_ptr, col, hdr)
            c.font = Font(bold=True, name="Arial", size=9)
            c.fill = COL_FILL
            c.alignment = Alignment(horizontal="center")
        row_ptr += 1

        total_vol = profile["total_volume"]
        poc, vah, val = profile["poc"], profile["vah"], profile["val"]

        for price, vol in sorted(zip(profile["bin_prices"], profile["bin_volumes"]), reverse=True):
            pct   = vol / total_vol * 100 if total_vol else 0
            level = ("POC" if abs(price - poc) < 1e-6 else
                     "VAH" if abs(price - vah) < 1e-6 else
                     "VAL" if abs(price - val) < 1e-6 else "")
            fill  = (POC_FILL if level == "POC" else
                     VAH_FILL if level == "VAH" else
                     VAL_FILL if level == "VAL" else None)

            ws.cell(row_ptr, 1, round(price, 2)).font = Font(name="Arial", size=9)
            ws.cell(row_ptr, 2, round(vol, 0)).font   = Font(name="Arial", size=9)
            ws.cell(row_ptr, 3, f"{pct:.1f}%").font   = Font(name="Arial", size=9)
            ws.cell(row_ptr, 4, level).font            = Font(bold=bool(level), name="Arial", size=9)
            for col in range(1, 5):
                ws.cell(row_ptr, col).alignment = Alignment(horizontal="center")
                if fill:
                    ws.cell(row_ptr, col).fill = fill
            row_ptr += 1

        row_ptr += 1  # blank separator

    for col, w in zip([1, 2, 3, 4], [12, 14, 14, 8]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A1"


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    wb           = Workbook()
    wb.remove(wb.active)
    summary_rows = []

    for symbol in SYMBOLS:
        print(f"Processing {symbol} ...", flush=True)
        df = fetch_daily_ohlcv(symbol, LOOKBACK_DAYS)
        if df.empty:
            print(f"  ⚠ No data, skipping.")
            continue

        sessions = []
        for date, row in df.iterrows():
            profile = build_session_profile(row)
            if not profile:
                continue
            date_str = date.strftime("%Y-%m-%d")
            sessions.append({"date": date_str, "profile": profile, "ohlcv": row})
            summary_rows.append([
                symbol, date_str,
                round(float(row["Open"]),  2),
                round(float(row["High"]),  2),
                round(float(row["Low"]),   2),
                round(float(row["Close"]), 2),
                int(float(row["Volume"])),
                round(profile["poc"], 2),
                round(profile["vah"], 2),
                round(profile["val"], 2),
                round(float(row["High"]) - float(row["Low"]), 2),
                round(profile["vah"] - profile["val"], 2),
            ])

        if sessions:
            write_symbol_sheet(wb, symbol, sessions)
            print(f"  ✓ {len(sessions)} sessions written")

    write_summary_sheet(wb, summary_rows)
    wb.save(DAILY_EXCEL)
    print(f"\n✅ Saved → {DAILY_EXCEL}  ({len(summary_rows)} session rows)")


if __name__ == "__main__":
    main()
