"""
add_sr_levels.py
================
Step 2 of 3  —  Cluster POC/VAH/VAL across sessions → Strong S/R zones.
                Annotates each symbol sheet and writes "SR Zones" summary.
                Pushes SR Zones tab to Google Sheets.

Reads:   volume_profile_daily.xlsx
Outputs: volume_profile_sr.xlsx
         Google Sheet tab → "SR Zones"

Run weekly on GitHub Actions after volume_profile.py.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import (
    SYMBOLS,
    DAILY_EXCEL, SR_EXCEL,
    DAILY_TOLERANCE_PCT, DAILY_MIN_TOUCHES, DAILY_STRONG_TOUCHES, SR_TOP_N,
    GOOGLE_SHEET_ID, GOOGLE_CREDS_JSON, SHEET_SR_ZONES,
)

# ── Colour palette ─────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F3864")
COL_FILL = PatternFill("solid", fgColor="D9E1F2")
SR_FILLS = {
    "STRONG_R": PatternFill("solid", fgColor="C00000"),
    "STRONG_S": PatternFill("solid", fgColor="375623"),
    "R":        PatternFill("solid", fgColor="FF9999"),
    "S":        PatternFill("solid", fgColor="C6EFCE"),
}
SR_FONTS = {
    "STRONG_R": Font(bold=True, color="FFFFFF", name="Arial", size=9),
    "STRONG_S": Font(bold=True, color="FFFFFF", name="Arial", size=9),
    "R":        Font(bold=True, color="C00000", name="Arial", size=9),
    "S":        Font(bold=True, color="375623", name="Arial", size=9),
}


# ── Parser ─────────────────────────────────────────────────────────────────────

def parse_symbol_sheet(ws) -> list:
    """
    Reads a symbol sheet written by volume_profile.py.
    Returns list of session dicts: {date, close, poc, vah, val, bin_rows}.
    bin_rows = [(row_number, price), ...]
    """
    sessions, current = [], None

    for r in range(1, ws.max_row + 1):
        v0 = ws.cell(r, 1).value
        v1 = ws.cell(r, 2).value

        if isinstance(v0, str) and v0.startswith("Date:"):
            if current:
                sessions.append(current)
            close = None
            if isinstance(v1, str):
                for part in v1.split():
                    if part.startswith("C:"):
                        try:
                            close = float(part[2:])
                        except ValueError:
                            pass
            current = {
                "date":     v0.replace("Date:", "").strip(),
                "close":    close,
                "poc":      None,
                "vah":      None,
                "val":      None,
                "bin_rows": [],
            }
            continue

        if v0 == "Price Bin" or v0 is None:
            continue

        if current and isinstance(v0, (int, float)):
            price = float(v0)
            level = ws.cell(r, 4).value
            current["bin_rows"].append((r, price))
            if level == "POC":
                current["poc"] = price
            elif level == "VAH":
                current["vah"] = price
            elif level == "VAL":
                current["val"] = price

    if current:
        sessions.append(current)
    return sessions


# ── Clustering ─────────────────────────────────────────────────────────────────

def cluster_levels(prices: list, tolerance_pct: float) -> list:
    """Group nearby prices into clusters. Returns list of {mean, count, prices}."""
    if not prices:
        return []
    prices = sorted(prices)
    groups, group = [], [prices[0]]
    for p in prices[1:]:
        if abs(p - sum(group) / len(group)) / (sum(group) / len(group)) * 100 <= tolerance_pct:
            group.append(p)
        else:
            groups.append(group)
            group = [p]
    groups.append(group)
    return [{"mean": sum(g) / len(g), "count": len(g), "prices": g} for g in groups]


def build_sr_zones(sessions: list, tolerance_pct: float, min_touches: int) -> list:
    """Collect POC/VAH/VAL → cluster → filter by min_touches → sort desc."""
    key_prices = [lvl for s in sessions
                  for lvl in [s["poc"], s["vah"], s["val"]] if lvl is not None]
    clusters = cluster_levels(key_prices, tolerance_pct)
    zones    = [c for c in clusters if c["count"] >= min_touches]
    zones.sort(key=lambda z: z["mean"], reverse=True)
    return zones


def pick_top_zones(zones: list, last_close: float, strong_touches: int, top_n: int):
    """Return (resistances[:top_n], supports[:top_n]) — STRONG only, nearest first."""
    strong = [z for z in zones if z["count"] >= strong_touches]
    res    = sorted([z for z in strong if z["mean"] >= last_close], key=lambda z: z["mean"])[:top_n]
    sup    = sorted([z for z in strong if z["mean"] <  last_close], key=lambda z: z["mean"], reverse=True)[:top_n]
    return res, sup


def sr_label(price, zones, last_close, strong_touches):
    for z in zones:
        if abs(price - z["mean"]) / z["mean"] * 100 <= DAILY_TOLERANCE_PCT:
            above  = price >= last_close if last_close else True
            strong = z["count"] >= strong_touches
            return ("STRONG_R" if strong and above else
                    "STRONG_S" if strong else
                    "R"        if above  else "S")
    return None


# ── Annotator ──────────────────────────────────────────────────────────────────

def annotate_symbol_sheet(ws, sessions: list, zones: list) -> None:
    """Write S/R label into column E for every bin row in the symbol sheet."""
    last_close = sessions[-1]["close"] if sessions else None
    row_map    = {rn: price for s in sessions for rn, price in s["bin_rows"]}

    for rn, price in row_map.items():
        label = sr_label(price, zones, last_close, DAILY_STRONG_TOUCHES)
        cell  = ws.cell(rn, 5)
        if label:
            cell.value     = label
            cell.fill      = SR_FILLS[label]
            cell.font      = SR_FONTS[label]
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.value = None

    # Column E header for every sub-table
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Price Bin":
            c           = ws.cell(r, 5)
            c.value     = "S/R"
            c.font      = Font(bold=True, name="Arial", size=9)
            c.fill      = COL_FILL
            c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["E"].width = 10


# ── Excel SR summary sheet ─────────────────────────────────────────────────────

def write_sr_summary_sheet(wb, sr_data: list) -> list:
    """
    Create/replace 'SR Zones' sheet.
    Returns export_rows (header + data) for Google Sheets push.
    """
    if "SR Zones" in wb.sheetnames:
        del wb["SR Zones"]
    ws = wb.create_sheet("SR Zones", 1)

    headers = ["Symbol", "Zone Price", "Touches", "Type", "Strength"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(1, col)
        c.fill = HDR_FILL
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.alignment = Alignment(horizontal="center")

    export_rows = [headers[:]]

    for symbol, zones, last_close in sr_data:
        res, sup = pick_top_zones(zones, last_close, DAILY_STRONG_TOUCHES, SR_TOP_N)
        for z in res + sup:
            price     = round(z["mean"], 2)
            zone_type = "Resistance" if price >= last_close else "Support"
            row_data  = [symbol, price, z["count"], zone_type, "STRONG"]
            ws.append(row_data)
            export_rows.append(row_data[:])
            r        = ws.max_row
            fill_key = "STRONG_R" if zone_type == "Resistance" else "STRONG_S"
            for col in range(1, 6):
                ws.cell(r, col).font      = Font(name="Arial", size=9)
                ws.cell(r, col).alignment = Alignment(horizontal="center")
            ws.cell(r, 2).fill = SR_FILLS[fill_key]
            ws.cell(r, 4).fill = SR_FILLS[fill_key]

    for i, w in enumerate([12, 12, 10, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return export_rows


# ── Google Sheets push ─────────────────────────────────────────────────────────

def push_sr_to_sheets(export_rows: list) -> None:
    """
    Push SR Zones data to Google Sheets tab SHEET_SR_ZONES.
    Column layout in Sheet:
      A Symbol | B Zone Price | C Touches | D Type | E Strength
      F LTP (Live GOOGLEFINANCE formula)
      G % from LTP (formula vs Zone Price)
      H Signal (formula: near zone? ⚠️ or ✅)
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = ["https://spreadsheets.google.com/feeds",
                  "https://www.googleapis.com/auth/drive"]
        creds  = Credentials.from_service_account_file(GOOGLE_CREDS_JSON, scopes=scopes)
        gc     = gspread.authorize(creds)
        sh     = gc.open_by_key(GOOGLE_SHEET_ID)

        try:
            ws = sh.worksheet(SHEET_SR_ZONES)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=SHEET_SR_ZONES, rows=300, cols=10)

        ws.clear()

        # ── Headers (row 1) ────────────────────────────────────────────────
        sheet_headers = ["Symbol", "Zone Price", "Touches", "Type", "Strength",
                         "LTP (Live)", "% from LTP", "Signal"]
        ws.update("A1", [sheet_headers], value_input_option="RAW")

        # ── Data rows with formulas ────────────────────────────────────────
        data_rows = export_rows[1:]   # skip header
        batch     = []

        for i, r in enumerate(data_rows, start=2):
            symbol, zone_price = r[0], float(r[1])
            gf_ticker          = f"NSE:{symbol}"

            ltp_formula    = f'=IFERROR(GOOGLEFINANCE("{gf_ticker}","price"),"")'
            pct_formula    = f'=IF(F{i}="","",ROUND((F{i}-B{i})/B{i}*100,2))'
            signal_formula = (
                f'=IF(F{i}="","⏳",IF(ABS(F{i}-B{i})/B{i}<=0.005,"⚠️ Near Zone","✅ Clear"))'
            )

            # Static columns A–E
            batch.append({"range": f"A{i}:E{i}",
                          "values": [[r[0], r[1], r[2], r[3], r[4]]]})
            # Formula columns F–H
            batch.append({"range": f"F{i}", "values": [[ltp_formula]]})
            batch.append({"range": f"G{i}", "values": [[pct_formula]]})
            batch.append({"range": f"H{i}", "values": [[signal_formula]]})

        ws.batch_update(batch, value_input_option="USER_ENTERED")

        # ── Formatting ────────────────────────────────────────────────────
        n = len(data_rows)
        requests = [
            # Header style
            {"repeatCell": {
                "range": {"sheetId": ws.id, "startRowIndex": 0, "endRowIndex": 1,
                          "startColumnIndex": 0, "endColumnIndex": 8},
                "cell": {"userEnteredFormat": {
                    "backgroundColor": {"red": 0.12, "green": 0.22, "blue": 0.39},
                    "textFormat": {"bold": True,
                                   "foregroundColor": {"red":1,"green":1,"blue":1}},
                    "horizontalAlignment": "CENTER",
                }},
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
            }},
            # Data centre-align
            {"repeatCell": {
                "range": {"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": n+1,
                          "startColumnIndex": 0, "endColumnIndex": 8},
                "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER"}},
                "fields": "userEnteredFormat.horizontalAlignment",
            }},
            # Signal ⚠️ → pink bg
            {"addConditionalFormatRule": {"rule": {
                "ranges": [{"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": n+1,
                            "startColumnIndex": 7, "endColumnIndex": 8}],
                "booleanRule": {
                    "condition": {"type": "TEXT_CONTAINS",
                                  "values": [{"userEnteredValue": "Near"}]},
                    "format": {"backgroundColor": {"red":1.0,"green":0.8,"blue":0.8}},
                },
            }, "index": 0}},
            # Type Resistance → red text
            {"addConditionalFormatRule": {"rule": {
                "ranges": [{"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": n+1,
                            "startColumnIndex": 3, "endColumnIndex": 4}],
                "booleanRule": {
                    "condition": {"type": "TEXT_EQ",
                                  "values": [{"userEnteredValue": "Resistance"}]},
                    "format": {"textFormat": {
                        "foregroundColor": {"red":0.75,"green":0,"blue":0}, "bold": True}},
                },
            }, "index": 1}},
            # Type Support → green text
            {"addConditionalFormatRule": {"rule": {
                "ranges": [{"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": n+1,
                            "startColumnIndex": 3, "endColumnIndex": 4}],
                "booleanRule": {
                    "condition": {"type": "TEXT_EQ",
                                  "values": [{"userEnteredValue": "Support"}]},
                    "format": {"textFormat": {
                        "foregroundColor": {"red":0.18,"green":0.49,"blue":0.19}, "bold": True}},
                },
            }, "index": 2}},
            # Freeze row 1
            {"updateSheetProperties": {
                "properties": {"sheetId": ws.id,
                               "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount",
            }},
        ]
        # Column widths
        for ci, px in enumerate([100, 100, 80, 110, 90, 100, 110, 120]):
            requests.append({"updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS",
                          "startIndex": ci, "endIndex": ci+1},
                "properties": {"pixelSize": px},
                "fields": "pixelSize",
            }})

        sh.batch_update({"requests": requests})
        print(f"✅ Google Sheets '{SHEET_SR_ZONES}' updated → {n} rows")

    except Exception as e:
        print(f"⚠ Google Sheets push failed: {e}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    wb           = openpyxl.load_workbook(DAILY_EXCEL)
    symbol_sheets = [s for s in wb.sheetnames
                     if s not in ("Summary", "SR Zones", "Weekly SR Zones")]
    sr_data = []

    for symbol in symbol_sheets:
        ws       = wb[symbol]
        sessions = parse_symbol_sheet(ws)
        if not sessions:
            print(f"⚠ {symbol}: no sessions parsed, skipping.")
            continue

        zones      = build_sr_zones(sessions, DAILY_TOLERANCE_PCT, DAILY_MIN_TOUCHES)
        last_close = sessions[-1]["close"]

        res, sup = pick_top_zones(zones, last_close, DAILY_STRONG_TOUCHES, SR_TOP_N)
        print(f"{symbol}: {len(res)} STRONG_R  {len(sup)} STRONG_S  (LTP ₹{last_close})")
        for z in res:
            print(f"  🔴 R  ₹{z['mean']:.2f}  ({z['count']} touches)")
        for z in sup:
            print(f"  🟢 S  ₹{z['mean']:.2f}  ({z['count']} touches)")

        annotate_symbol_sheet(ws, sessions, zones)
        sr_data.append((symbol, zones, last_close))

    export_rows = write_sr_summary_sheet(wb, sr_data)
    wb.save(SR_EXCEL)
    print(f"\n✅ Saved → {SR_EXCEL}")

    if GOOGLE_SHEET_ID:
        push_sr_to_sheets(export_rows)


if __name__ == "__main__":
    main()
